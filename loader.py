"""엑셀 로더: universities.xlsx → 대학 레코드 목록 + 링크 수집.

- 헤더 행(6)에서 A~X 컬럼을 config.COL 매핑으로 정규화
- 셀 내장 하이퍼링크 + 셀 텍스트 안의 URL을 모두 수집
- 국가 필터 지원
"""
import re
import openpyxl
from openpyxl.utils import column_index_from_string
import config

URL_RE = re.compile(r"https?://[^\s)\]\">,]+")


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _collect_links(cell_value_map, ws, row):
    """행의 지정 컬럼들에서 링크를 우선순위대로 수집(중복 제거)."""
    links = []
    seen = set()

    def add(url, col_key):
        url = url.strip().rstrip(".,;")
        if not url or url in seen:
            return
        seen.add(url)
        links.append({"url": url, "col": col_key})

    for key in config.LINK_COLUMNS:
        col_letter = config.COL[key]
        ci = column_index_from_string(col_letter)
        cell = ws.cell(row=row, column=ci)
        # 1) 셀에 내장된 하이퍼링크
        if cell.hyperlink and cell.hyperlink.target:
            add(cell.hyperlink.target, key)
        # 2) 셀 텍스트 안의 URL
        txt = cell_value_map.get(key)
        if txt:
            for m in URL_RE.findall(txt):
                add(m, key)
    return links


def load(xlsx_path=None, country=None, limit=None):
    xlsx_path = xlsx_path or config.DEFAULT_XLSX
    wb_val = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_link = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws_val = wb_val.worksheets[0]
    ws_link = wb_link.worksheets[0]

    records = []
    for row in range(config.FIRST_DATA_ROW, ws_val.max_row + 1):
        # 값 읽기
        vals = {}
        for key, col_letter in config.COL.items():
            ci = column_index_from_string(col_letter)
            vals[key] = _clean(ws_val.cell(row=row, column=ci).value)

        if not vals.get("country") and not vals.get("university"):
            continue
        if country and (vals.get("country") or "").strip().upper() != country.strip().upper():
            continue

        links = _collect_links(vals, ws_link, row)
        rec = {
            "row": row,
            "country": vals.get("country"),
            "university": vals.get("university"),
            "sheet": vals,          # 원본 셀 값 전부 보존
            "links": links,
        }
        records.append(rec)
        if limit and len(records) >= limit:
            break
    return records


if __name__ == "__main__":
    import sys, json
    c = sys.argv[1] if len(sys.argv) > 1 else "UNITED STATES"
    recs = load(country=c)
    print(f"{c}: {len(recs)} universities")
    for r in recs[:3]:
        print("-", r["university"], "| links:", len(r["links"]))
        for l in r["links"]:
            print("    ", l["col"], l["url"][:90])

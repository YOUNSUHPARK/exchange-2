"""data.json → 가독성 있는 CSV + 엑셀 표. 값은 깨끗하게, 출처는 별도 요약 컬럼."""
import csv
import json
import os
import config

COLUMNS = [
    ("국가", lambda r: r["country"]),
    ("대학명", lambda r: r["university"]),
    ("지원가능 전공(SKKU)", lambda r: r.get("major_skku")),
    ("해외대학 수학 전공", lambda r: v(r["major_abroad"])),
    ("최소 GPA", lambda r: v(r["gpa_min"])),
    ("완료 학기 수", lambda r: v(r.get("semesters"))),
    ("최소 학점", lambda r: v(r.get("credits_min"))),
    ("최대 학점", lambda r: v(r.get("credits_max"))),
    ("위치(도심 거리)", lambda r: v(r["location"])),
    ("수업 언어", lambda r: v(r["lang_instruction"])),
    ("어학 요건", lambda r: v(r["lang_req"])),
    ("출처 요약", lambda r: sources_summary(r)),
    ("참고 링크", lambda r: r.get("program_link")),
    ("문의 이메일", lambda r: r.get("email")),
]


def v(obj):
    return obj.get("value") if isinstance(obj, dict) else obj


def sources_summary(r):
    parts = []
    for label, key in [("GPA", "gpa_min"), ("어학", "lang_req"), ("수업언어", "lang_instruction"),
                       ("전공", "major_abroad"), ("위치", "location")]:
        s = r[key].get("source") if isinstance(r.get(key), dict) else None
        if s:
            parts.append(f"{label}:{s}")
    return " · ".join(parts)


def build(records, out_dir=None):
    out_dir = out_dir or config.OUT_DIR
    os.makedirs(out_dir, exist_ok=True)
    headers = [c[0] for c in COLUMNS]

    csv_path = os.path.join(out_dir, "universities.csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in records:
            w.writerow([c[1](r) or "" for c in COLUMNS])

    # 엑셀
    xlsx_path = None
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "universities"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5496")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for r in records:
            ws.append([c[1](r) or "" for c in COLUMNS])
        widths = [12, 26, 16, 30, 14, 10, 12, 12, 26, 10, 26, 24, 30, 24]
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wdt
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        xlsx_path = os.path.join(out_dir, "universities.xlsx")
        wb.save(xlsx_path)
    except Exception as e:
        print(f"[table] 엑셀 생성 생략: {e}")

    print(f"[table] CSV → {csv_path}")
    if xlsx_path:
        print(f"[table] XLSX → {xlsx_path}")
    return csv_path, xlsx_path


if __name__ == "__main__":
    data = json.load(open(os.path.join(config.HERE, "data.json"), encoding="utf-8"))
    build(data)
